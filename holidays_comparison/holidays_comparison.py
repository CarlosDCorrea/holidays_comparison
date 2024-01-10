import sys
from datetime import datetime, date

from openpyxl import Workbook, load_workbook

import holidays

from .constants import (
    COLOMBIA_TITLE_CELL,
    COLOMBIA_TOTAL_TITLE_CELL,
    CHILE_TITLE_CELL,
    CHILE_TOTAL_TITLE_CELL,
    DATE_TITLE_CELL,
    START_CELL,
    RED_COLOR,
    GREEN_COLOR,
    FILE_NAME
)


# we only need week holidays so we filter the ones that are monday-friday
def _is_weekend(date: datetime) -> bool:
    return date.weekday() in range(5, 7)


def _create_wb():
    """ Either creates or returns a Workbook """
    try:
        return load_workbook(FILE_NAME)
    except FileNotFoundError:
        print(f'The file named \'{FILE_NAME}\' has not been found, creating a new one')
        return Workbook()
    except Exception as e:
        print(f'Some other exception has ocurred while reading the file {e}')
        sys.exit(1)


def _create_worksheet(wb, year):
    ws = wb.active

    if ws.title == 'Sheet':
        ws.title = str(year)
        return ws
    elif ws.title == str(year) or str(year) in wb.sheetnames:
        print(f'Another sheet with the same year {year} already exist, exiting app')
        sys.exit(1)
    else:
        return wb.create_sheet(title=str(year))


def perform_comparison(args):
    year: int = args.y if args.y else date.today().year

    # get colombia and chile holidays as dictionaries
    colombia_holidays: dict[datetime, str] = holidays.CO(years=year)
    chile_holidays: dict[datetime, str] = holidays.CL(years=year)

    # transform both holidays into sets to delete duplications
    colombia_holidays_set = {
        date_ for date_ in colombia_holidays.keys() if not _is_weekend(date_)}
    chile_holidays_set = {
        date_ for date_ in chile_holidays.keys() if not _is_weekend(date_)}
    holidays_set = colombia_holidays_set | chile_holidays_set

    # since sets are not subscriptable, we need to get sorted list from them
    holidays_list = sorted(list(holidays_set),
                           key=lambda date: datetime.strptime(date.strftime('%d/%m/%y'), '%d/%m/%y'))

    AMOUNT_OF_HOLIDAYS = len(holidays_set)

    wb = _create_wb()

    ws = _create_worksheet(wb, year)

    ws[DATE_TITLE_CELL] = 'Fecha'
    ws[COLOMBIA_TITLE_CELL] = 'Colombia'
    ws[CHILE_TITLE_CELL] = 'Chile'

    # define rows, the numbers of rows will be the number of holidays in the specified year
    date_rows = ws[f'G{START_CELL}':f'G{START_CELL + AMOUNT_OF_HOLIDAYS - 1}']
    colombia_rows = ws[f'H{START_CELL}':f'H{START_CELL + AMOUNT_OF_HOLIDAYS - 1}']
    chile_rows = ws[f'I{START_CELL}':f'I{START_CELL + AMOUNT_OF_HOLIDAYS - 1}']

    ws[COLOMBIA_TOTAL_TITLE_CELL] = 'Totales Colombia'
    ws[CHILE_TOTAL_TITLE_CELL] = 'Totales Chile'

    ws['K11'] = len(colombia_holidays_set)
    ws['L11'] = len(chile_holidays_set)

    for i in range(len(date_rows)):
        date_rows[i][0].value = holidays_list[i]

        colombia_rows[i][0].fill = GREEN_COLOR if holidays_list[i] in colombia_holidays_set else RED_COLOR
        chile_rows[i][0].fill = GREEN_COLOR if holidays_list[i] in chile_holidays_set else RED_COLOR

        colombia_rows[i][0].value = 'Festivo' if holidays_list[i] in colombia_holidays_set else 'No Festivo'
        chile_rows[i][0].value = 'Festivo' if holidays_list[i] in chile_holidays_set else 'No Festivo'

    wb.save(FILE_NAME)
